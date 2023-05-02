import os
import openpyxl

path = "C:\\Data Analysis Automation\\Data Files"
ewr_database_path = "C:\\Data Analysis Automation\\EWR Database.xlsx"

files = os.listdir(path)
file_directory = []
for file in files:
    file_directory.append(path + "\\" + file)

data_path = file_directory[0]
print(data_path)
wb = openpyxl.load_workbook(data_path)

ewr_wb = openpyxl.load_workbook(ewr_database_path)
ewr_ws = ewr_wb.active
ewr_ws.title = "BCIS-06"
ewr_ws_location = ewr_ws["A1"]
ewr_ws_location.value = 100000
ewr_wb.save(ewr_database_path)


for sheet in wb.sheetnames:
    if sheet == "Check In":
        ws = wb[sheet]



